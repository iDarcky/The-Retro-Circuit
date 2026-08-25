import openpyxl, json, re, datetime
SRC='/root/.claude/uploads/454814d1-9ce8-55b6-9119-8737128283a8/8c683539-handhelds_v2.xlsx'
ws=openpyxl.load_workbook(SRC, data_only=True)['Sheet1']

# Emulation columns read from the sheet's OWN headers (order differs from v1; adds PS3)
EMU={7:['gb_state','gbc_state'],8:['nes_state'],9:['genesis_state'],10:['gba_state'],11:['snes_state'],
     12:['ps1_state'],13:['nds_state'],14:['n64_state'],15:['dreamcast_state'],16:['psp_state'],
     17:['saturn_state'],18:['gamecube_state'],19:['x3ds_state'],20:['wii_state'],21:['ps2_state'],
     22:['switch_state'],23:['wii_u'],24:['ps3_state']}
GRADE={'A':'Perfect','A-':'Great','B+':'Great','B':'Great','B-':'Playable','C+':'Playable',
       'C':'Playable','C-':'Struggles','C--':'Struggles','D+':'Struggles','D':'Struggles',
       'D-':'Struggles','F':'Unplayable'}
NULLISH={'❌','?','-','n/a','N/A','—','','None','?)'}

def cell(r,c):
    v=ws.cell(row=r,column=c).value
    if v is None: return None
    if isinstance(v,(datetime.time,datetime.timedelta,datetime.datetime)): return v
    s=' '.join(str(v).split()) if '\n' not in str(v) else str(v)
    return None if s.strip() in NULLISH else (s.strip() or None)

def first(s):
    if s is None: return None
    v=str(s).split('\n')[0].strip().rstrip(',;').strip()
    return None if v in NULLISH else (v or None)
def flat(s):
    if s is None: return None
    v=' '.join(x.strip() for x in str(s).split('\n') if x.strip()).strip().rstrip(',;').strip()
    return None if v in NULLISH else (v or None)
def num(s):
    if s is None: return None
    m=re.search(r'(\d+(?:\.\d+)?)', str(s).replace(',',''))
    return float(m.group(1)) if m else None
def mhz(s):
    if s is None: return None
    best=None
    for val,unit in re.findall(r'(\d+(?:\.\d+)?)\s*(GHz|MHz)', str(s), re.I):
        v=float(val)*(1000 if unit.lower()=='ghz' else 1)
        best=v if best is None else max(best,v)
    return int(round(best)) if best else None
def ram_mb(s):
    if s is None: return None
    m=re.search(r'(\d+(?:\.\d+)?)\s*(GB|MB)', str(s), re.I)   # first = base config
    if not m: return None
    v=float(m.group(1)); return int(v*1024) if m.group(2).upper()=='GB' else int(v)
def cores(s):
    if s is None: return None
    low=str(s).lower()
    for w,n in {'single':1,'dual':2,'triple':3,'quad':4,'hexa':6,'octa':8}.items():
        if w in low: return n
    m=re.search(r'(\d+)\s*(core|thread)', low)
    return int(m.group(1)) if m else None
def resolution(s):
    if s is None: return (None,None)
    m=re.search(r'(\d+)\s*[×x]\s*(\d+)', str(s))
    return (int(m.group(1)), int(m.group(2))) if m else (None,None)
def dims(s):
    if s is None: return (None,None,None)
    m=re.findall(r'(\d+(?:\.\d+)?)\s*mm', str(s))
    return (float(m[0]),float(m[1]),float(m[2])) if len(m)>=3 else (None,None,None)
def release(s):
    if s is None: return (None,None)
    t=str(s)
    m=re.match(r'(\d{4})\s*/\s*(\d{1,2})', t)
    if m: return ("%s-%02d-01"%(m.group(1),int(m.group(2))),'month')
    m=re.match(r'(\d{4})', t)
    return ("%s-01-01"%m.group(1),'year') if m else (None,None)
def aspect(v):
    # Excel silently converted "16:9" into a TIME value (16:09:00). Rebuild the ratio.
    if isinstance(v, datetime.time): return "%d:%d"%(v.hour, v.minute)
    if isinstance(v,(datetime.timedelta,datetime.datetime)): return None
    return first(v)
def os_bits(s):
    if not s: return (None,None)
    low=str(s).lower()
    fam=('android' if 'android' in low or 'andorid' in low else
         'windows' if 'windows' in low else 'steamos' if 'steamos' in low else
         'linux' if 'linux' in low else 'proprietary')
    m=re.search(r'(\d+(?:\.\d+)*)', str(s))
    return (fam, m.group(1) if m else None)
def form_factor(s):
    if not s: return None
    low=str(s).lower()
    if 'clamshell' in low: return 'Clamshell'
    if 'vertical' in low: return 'Vertical'
    if 'horizontal' in low or 'micro horizontal' in low: return 'Horizontal'
    return 'Other'
def slugify(s):
    # Preserve '+' as '-plus' so e.g. "GPD XD+" and "GPD XD" don't collapse to the same slug.
    t=str(s).lower().strip().replace('+',' plus ')
    return re.sub(r'(^-|-$)','', re.sub(r'[^a-z0-9]+','-', t))

# Brand normalisation: take the primary company, drop parentheticals/co-marketing partners.
BRAND={'anbernic':'Anbernic','retroid / moorechip':'Retroid','ayaneo':'Ayaneo',
 'ayn technologies':'Ayn','powkiddy':'Powkiddy','trimui':'Trimui','kinhank':'KinHank',
 'konkr (ayaneo)':'KONKR','kt pocket':'KT Pocket','mangmi':'Mangmi','miyoo':'Miyoo',
 'razer, verizon, qualcomm':'Razer','logitech, tencent':'Logitech','serious play / n+':'Serious Play',
 'yinlips / smaggi':'Yinlips','szdiier / diium':'SZDIIER','snail / ireadygo / 78dian':'Snail',
 'onexsugar (one netbook & sugar cube)':'ONEXSUGAR','campaign (?)':'Campaign','"m god"':'M God',
 '乐天游':'Lotiyo','gamepad digital':'GPD','jinxing digital':'JinXing Digital'}
def brand_of(b):
    b=' '.join(str(b or '').split())
    return BRAND.get(b.lower(), b)

dedup=json.load(open('v2_dedup.json'))
skip={(b,n) for b,n in dedup['dupes']}

out=[]
for r in range(2, ws.max_row+1):
    name=flat(cell(r,1)); braw=' '.join(str(ws.cell(row=r,column=2).value or '').split())
    if not name or not braw: continue
    if (braw,name) in skip: continue
    brand=brand_of(braw)
    # strip a leading brand prefix so names read like the existing catalogue ("Pocket 4")
    disp=re.sub(r'^%s\s+'%re.escape(brand), '', name, flags=re.I).strip() or name
    rel,prec=release(cell(r,3))
    osr=first(cell(r,5)); fam,ver=os_bits(osr)
    rx,ry=resolution(cell(r,38)); w,h,dp=dims(cell(r,60))
    price=cell(r,70); pv=None
    if price and 'discontinu' not in str(price).lower():
        m=re.search(r'\$\s*(\d+)', str(price)) or re.search(r'(\d+)', str(price))
        pv=int(m.group(1)) if m else None
    st=flat(cell(r,50))
    v={'variant_name':'Base','is_default':True,'release_date':rel,'release_date_precision':prec,
       'price_launch_usd':pv,'os':osr,'os_family':fam,'os_version':ver,
       'soc':first(cell(r,25)),'cpu_model':flat(cell(r,26)),'cpu_cores':cores(cell(r,27)),
       'cpu_threads':cores(cell(r,28)),'cpu_clock_max_mhz':mhz(cell(r,29)),
       'cpu_architecture':first(cell(r,30)),'gpu_model':flat(cell(r,31)),
       'gpu_cores':cores(cell(r,32)),'gpu_clock_mhz':mhz(cell(r,33)),'ram_mb':ram_mb(cell(r,34)),
       'screen_size_inch':num(cell(r,35)),'display_type':first(cell(r,36)),
       'refresh_rate_hz':int(num(cell(r,37))) if num(cell(r,37)) else None,
       'screen_resolution_x':rx,'screen_resolution_y':ry,
       'ppi':int(num(cell(r,39))) if num(cell(r,39)) else None,
       'aspect_ratio':aspect(ws.cell(row=r,column=40).value),
       'battery_capacity_mah':int(num(cell(r,42))) if num(cell(r,42)) else None,
       'cooling_solution':flat(cell(r,43)),'storage_type':st,
       'storage_expandable': bool(st and ('external' in st.lower() or 'microsd' in st.lower())),
       'other_connectivity':flat(cell(r,51)),'video_out':flat(cell(r,52)),
       'weight_g':int(num(cell(r,61))) if num(cell(r,61)) else None,
       'body_material':first(cell(r,62)),'available_colors':flat(cell(r,63)),
       'width_mm':w,'height_mm':h,'depth_mm':dp,
       'touchscreen': bool(cell(r,36) and 'touch' in str(cell(r,36)).lower())}
    v={k:x for k,x in v.items() if x is not None}
    emu={}
    for col,fields in EMU.items():
        raw=cell(r,col)
        if not raw: continue
        g=str(raw).split('\n')[-1].strip()
        s=GRADE.get(g)
        if s:
            for f in fields: emu[f]=s
    if emu: v['emulation']=emu
    out.append({'name':disp,'manufacturer':brand,'slug':slugify(brand+' '+disp),
                'form_factor':form_factor(cell(r,4)),'device_category':'emulation',
                'status':'draft','variants':[v]})

json.dump(out, open('v2_import.json','w'), indent=1, ensure_ascii=False)
slugs=[o['slug'] for o in out]
print("consoles:", len(out), "| unique slugs:", len(set(slugs)))
dups=[s for s in set(slugs) if slugs.count(s)>1]
print("DUPLICATE SLUGS:", dups)
print("emulation entries:", sum(len(o['variants'][0].get('emulation',{})) for o in out))
print("brands:", len({o['manufacturer'] for o in out}))

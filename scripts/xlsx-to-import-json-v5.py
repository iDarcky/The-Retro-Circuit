"""
xlsx-to-import-json-v5.py
-------------------------
Full-fidelity converter for the handhelds spreadsheets (v3 / v4 layout).

Supersedes v2/v3/v4, which read only ~30 of the 62 non-emulation columns. Everything
those dropped — buttons, speakers, audio, rumble, sensors, charge port, screen lens,
reviews, vendor links, pros/cons, notes — is mapped here. See docs/DATA_MAPPING.md for
the column-by-column reference this implements.

Output shape per console:
    {
      name, manufacturer, slug, form_factor, device_category, status,
      release_status, price_tier, pros[], cons[], description,
      links: [{kind, url, label, sort_order}],
      variants: [{ ...specs, emulation: {}, input_profile: {} }]
    }

Usage:
    python3 scripts/xlsx-to-import-json-v5.py <source.xlsx> <out.json> [--category emulation|pc_gaming]
"""

import openpyxl, json, re, sys, datetime

SRC = sys.argv[1]
OUT = sys.argv[2]
CATEGORY = 'pc_gaming' if '--category' in sys.argv and 'pc_gaming' in sys.argv else 'emulation'

ws = openpyxl.load_workbook(SRC, data_only=True)['Sheet1']

# Emulation columns, 1-indexed, read from the sheet's own header order.
EMU = {7: ['gb_state', 'gbc_state'], 8: ['nes_state'], 9: ['genesis_state'], 10: ['gba_state'],
       11: ['snes_state'], 12: ['ps1_state'], 13: ['nds_state'], 14: ['n64_state'],
       15: ['dreamcast_state'], 16: ['psp_state'], 17: ['saturn_state'], 18: ['gamecube_state'],
       19: ['x3ds_state'], 20: ['wii_state'], 21: ['ps2_state'], 22: ['switch_state'],
       23: ['wii_u'], 24: ['ps3_state']}
GRADE = {'A': 'Perfect', 'A-': 'Great', 'B+': 'Great', 'B': 'Great', 'B-': 'Playable',
         'C+': 'Playable', 'C': 'Playable', 'C-': 'Struggles', 'C--': 'Struggles',
         'D+': 'Struggles', 'D': 'Struggles', 'D-': 'Struggles', 'F': 'Unplayable'}
NULLISH = {'❌', '?', '-', 'n/a', 'N/A', '—', '', 'None', '?)'}


# ---------------------------------------------------------------- primitives
def raw(r, c):
    return ws.cell(row=r, column=c).value

def cell(r, c):
    v = raw(r, c)
    if v is None:
        return None
    if isinstance(v, (datetime.time, datetime.timedelta, datetime.datetime)):
        return v
    s = ' '.join(str(v).split()) if '\n' not in str(v) else str(v)
    return None if s.strip() in NULLISH else (s.strip() or None)

def first(s):
    if s is None: return None
    v = str(s).split('\n')[0].strip().rstrip(',;').strip()
    return None if v in NULLISH else (v or None)

def flat(s):
    if s is None: return None
    v = ' '.join(x.strip() for x in str(s).split('\n') if x.strip()).strip().rstrip(',;').strip()
    return None if v in NULLISH else (v or None)

def num(s):
    if s is None: return None
    m = re.search(r'(\d+(?:\.\d+)?)', str(s).replace(',', ''))
    return float(m.group(1)) if m else None

def mhz(s):
    if s is None: return None
    best = None
    for val, unit in re.findall(r'(\d+(?:\.\d+)?)\s*(GHz|MHz)', str(s), re.I):
        v = float(val) * (1000 if unit.lower() == 'ghz' else 1)
        best = v if best is None else max(best, v)
    return int(round(best)) if best else None

def ram_mb(s):
    if s is None: return None
    m = re.search(r'(\d+(?:\.\d+)?)\s*(GB|MB)', str(s), re.I)
    if not m: return None
    v = float(m.group(1))
    return int(v * 1024) if m.group(2).upper() == 'GB' else int(v)

def cores(s):
    if s is None: return None
    low = str(s).lower()
    for w, n in {'single': 1, 'dual': 2, 'triple': 3, 'quad': 4, 'hexa': 6, 'octa': 8}.items():
        if w in low: return n
    m = re.search(r'(\d+)\s*(core|thread)', low)
    return int(m.group(1)) if m else None

def resolution(s):
    if s is None: return (None, None)
    m = re.search(r'(\d+)\s*[×x]\s*(\d+)', str(s))
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)

def dims(s):
    if s is None: return (None, None, None)
    m = re.findall(r'(\d+(?:\.\d+)?)\s*mm', str(s))
    return (float(m[0]), float(m[1]), float(m[2])) if len(m) >= 3 else (None, None, None)

def release(s):
    if s is None: return (None, None)
    t = str(s)
    m = re.match(r'(\d{4})\s*/\s*(\d{1,2})', t)
    if m: return ("%s-%02d-01" % (m.group(1), int(m.group(2))), 'month')
    m = re.match(r'(\d{4})', t)
    return ("%s-01-01" % m.group(1), 'year') if m else (None, None)

def aspect(v):
    # Excel silently turns "16:9" into a TIME value (16:09:00). Rebuild the ratio.
    if isinstance(v, datetime.time): return "%d:%d" % (v.hour, v.minute)
    if isinstance(v, (datetime.timedelta, datetime.datetime)): return None
    return first(v)

def tick(r, c):
    """A ✅/❌ column: True, False, or None when genuinely unknown ('?')."""
    v = raw(r, c)
    if v is None: return None
    s = str(v).strip()
    if '✅' in s or s.lower() in ('yes', 'true'): return True
    if '❌' in s or s.lower() in ('no', 'false'): return False
    return None


# ------------------------------------------------------- column-specific parsing
def os_bits(s):
    if not s: return (None, None)
    low = str(s).lower()
    fam = ('android' if 'android' in low or 'andorid' in low else
           'windows' if 'windows' in low else 'steamos' if 'steamos' in low else
           'linux' if 'linux' in low else 'proprietary')
    m = re.search(r'(\d+(?:\.\d+)*)', str(s))
    return (fam, m.group(1) if m else None)

def cpu_arch_enum(arch_text, cpu_text):
    """Map the sheet's free-text architecture onto the strict cpu_arch enum."""
    t = f"{arch_text or ''} {cpu_text or ''}".lower()
    if 'x86' in t or 'amd' in t or 'intel' in t or 'zen' in t or 'ryzen' in t: return 'x86_64'
    if 'mips' in t or 'xtensa' in t or 'risc' in t: return 'other'
    if 'arm' in t or 'cortex' in t:
        # 32-bit ARM cores that predate ARMv8.
        if re.search(r'arm(9|11|926)|cortex-a(5|7|8|9)\b', t): return 'arm32'
        return 'arm64'
    return None

def split_connectivity(s):
    """"WiFi 5, Bluetooth 4.2, USB-C OTG" -> (wifi, bluetooth, rest)."""
    if not s: return (None, None, None)
    wifi = bt = None
    rest = []
    for part in re.split(r',(?![^()]*\))', str(s)):
        p = part.strip()
        if not p: continue
        low = p.lower()
        if low.startswith('wifi') or low.startswith('wi-fi'):
            wifi = p if wifi is None else f"{wifi}, {p}"
        elif low.startswith('bluetooth'):
            bt = p if bt is None else f"{bt}, {p}"
        else:
            rest.append(p)
    clean = lambda v: None if v is None or '#?' in v else v
    return (clean(wifi), clean(bt), ', '.join(rest) or None)

def parse_storage(s):
    """"Internal 8 GB eMMC, Dual External MicroSD" -> (gb, type, expandable, microsd_type)."""
    if not s: return (None, None, None, None)
    t = str(s)
    low = t.lower()
    gb = None
    m = re.search(r'internal\s+(\d+(?:\.\d+)?)\s*(gb|tb|mb)', low)
    if m:
        v = float(m.group(1))
        gb = int(v * 1024) if m.group(2) == 'tb' else (v if m.group(2) == 'gb' else round(v / 1024, 2))
        gb = int(gb) if float(gb).is_integer() else gb
    stype = None
    for token in ('emmc', 'nvme', 'ssd', 'emcp', 'flash', 'microsd'):
        if token in low:
            stype = {'emmc': 'eMMC', 'nvme': 'NVMe SSD', 'ssd': 'SSD',
                     'emcp': 'eMCP', 'flash': 'Flash', 'microsd': 'MicroSD'}[token]
            break
    expandable = ('external' in low) or ('microsd' in low)
    microsd = None
    mu = re.search(r'(uhs-?i{1,2}|uhs-?[12])', low)
    if mu: microsd = mu.group(1).upper().replace('UHS', 'UHS-').replace('--', '-')
    elif 'dual external microsd' in low: microsd = 'Dual MicroSD'
    elif 'microsd' in low: microsd = 'MicroSD'
    return (gb, stype, expandable, microsd)

def parse_dpad(s):
    if not s: return (None, None)
    shape = None
    for k in ('Cross', 'Disc', 'Split', 'Segmented', 'Rocker'):
        if k.lower() in str(s).lower(): shape = k; break
    place = None
    m = re.search(r'(upper|lower|middle)\s+placement', str(s), re.I)
    if m: place = m.group(1).title()
    return (shape, place)

def parse_analogs(s):
    """"Dual thumbsticks (L3/R3, Hall) Lower placement" -> count, tech, clicks, layout."""
    if not s: return (0, None, None, None)
    t = str(s); low = t.lower()
    count = 2 if 'dual' in low else 1 if 'single' in low else 0
    tech = None
    for k in ('Hall', 'TMR', 'Potentiometer', 'ALPS'):
        if k.lower() in low: tech = k; break
    clicks = ('l3' in low or 'r3' in low)
    layout = None
    m = re.search(r'(upper|lower|middle|symmetric|asymmetric)', t, re.I)
    if m: layout = m.group(1).title()
    return (count, tech, clicks, layout)

def parse_shoulder(s):
    """Returns (bumper_tech, trigger_tech, layout)."""
    if not s: return (None, None, None)
    t = str(s); low = t.lower()
    has_l2 = 'l2' in low
    layout = None
    m = re.search(r'(horizontal|vertical|stacked|shelf)', t, re.I)
    if m: layout = m.group(1).title()
    trigger = None
    if 'analog' in low or 'hall' in low: trigger = 'Analog'
    elif has_l2: trigger = 'Digital'
    bumper = 'Digital' if 'l1' in low else None
    return (bumper, trigger, layout)

def stars(s):
    """"⭐️⭐️⭐️⭐️⭐️¼" -> "4.25/5"; "💥💥" -> "2/5 (estimate)"."""
    if not s: return None
    t = str(s)
    n = t.count('⭐') or t.count('💥')
    if not n: return flat(s)
    frac = 0.25 if '¼' in t else 0.5 if '½' in t else 0.75 if '¾' in t else 0
    est = ' (estimate)' if 'estimate' in t.lower() else ''
    total = n + frac
    return f"{total:g}/5{est}"

def slugify(s):
    t = str(s).lower().strip().replace('+', ' plus ')
    return re.sub(r'(^-|-$)', '', re.sub(r'[^a-z0-9]+', '-', t))


BRAND = {'anbernic': 'Anbernic', 'retroid / moorechip': 'Retroid', 'ayaneo': 'Ayaneo',
 'ayn technologies': 'Ayn', 'powkiddy': 'Powkiddy', 'trimui': 'Trimui', 'kinhank': 'KinHank',
 'konkr (ayaneo)': 'KONKR', 'kt pocket': 'KT Pocket', 'mangmi': 'Mangmi', 'miyoo': 'Miyoo',
 'razer, verizon, qualcomm': 'Razer', 'logitech, tencent': 'Logitech',
 'serious play / n+': 'Serious Play', 'yinlips / smaggi': 'Yinlips', 'szdiier / diium': 'SZDIIER',
 'snail / ireadygo / 78dian': 'Snail', 'onexsugar (one netbook & sugar cube)': 'ONEXSUGAR',
 'campaign (?)': 'Campaign', '"m god"': 'M God', '乐天游': 'Lotiyo', 'gamepad digital': 'GPD',
 'game pad digital': 'GPD', 'jinxing digital': 'JinXing Digital', 'miyoo / bittboy': 'Miyoo',
 'game kiddy': 'Game Kiddy', 'waveshare': 'Waveshare', 'hardkernel': 'HardKernel',
 'clockworkpi': 'ClockworkPi', 'retroflag': 'Retroflag', 'gamepark holdings': 'GamePark Holdings',
 'datafrog': 'Data Frog', 'bittboy / miyoo / wolsen': 'Miyoo',
 'bittboy / miyoo / wolsen (retromimi)': 'Miyoo', 'batlexp (anbernic?)': 'BATLEXP',
 'powkiddy, coolbaby': 'Powkiddy', 'subor, coolbaby': 'Subor', 'trimui / powkiddy': 'Trimui',
 'game kiddy / z-pocket game': 'Game Kiddy', 'ldk / wolsen': 'LDK',
 'dingoo digital technology': 'Dingoo', 'dingoo technology': 'Dingoo',
 'game consoles worldwide': 'Game Console', 'openpandora gmbh': 'OpenPandora',
 'one netbook': 'One Netbook', 'one netbook, tencent': 'One Netbook',
 'aokzoe (one netbook spinoff)': 'AOKZOE', 'asus & microsoft': 'Asus', 'msi': 'MSI',
 'lenovo': 'Lenovo', 'acer': 'Acer'}
DROP_BRAND = {'DIY', 'Unknown', 'ChangLiang Li', 'None'}

def brand_of(b):
    b = ' '.join(str(b or '').split())
    return BRAND.get(b.lower(), b)


# ---------------------------------------------------------------------- build
out, skipped = [], []
for r in range(2, ws.max_row + 1):
    name = flat(cell(r, 1))
    braw = ' '.join(str(raw(r, 2) or '').split())
    if not name or not braw:
        continue
    brand = brand_of(braw)
    if brand in DROP_BRAND:
        skipped.append(name); continue

    disp = re.sub(r'^%s\s+' % re.escape(brand), '', name, flags=re.I).strip() or name
    slug = slugify(brand + ' ' + disp)

    rel, prec = release(cell(r, 3))
    osr = first(cell(r, 5)); fam, ver = os_bits(osr)
    rx, ry = resolution(cell(r, 38))
    w, h, dp = dims(cell(r, 60))
    wifi, bt, other_conn = split_connectivity(flat(cell(r, 51)))
    sgb, stype, sexp, microsd = parse_storage(flat(cell(r, 50)))
    arch_text = first(cell(r, 30))
    cpu_text = flat(cell(r, 26))

    price = cell(r, 70); pv = None
    if price and 'discontinu' not in str(price).lower():
        m = re.search(r'\$\s*(\d+)', str(price)) or re.search(r'(\d+)', str(price))
        pv = int(m.group(1)) if m else None

    audio_out = flat(cell(r, 53))
    sensors = flat(cell(r, 56))

    v = {
        'variant_name': 'Base', 'is_default': True,
        'release_date': rel, 'release_date_precision': prec,
        'price_avg_usd': pv,
        'os': osr, 'os_family': fam, 'os_version': ver,
        'soc': first(cell(r, 25)), 'cpu_model': cpu_text,
        'cpu_cores': cores(cell(r, 27)), 'cpu_threads': cores(cell(r, 28)),
        'cpu_clock_max_mhz': mhz(cell(r, 29)),
        'cpu_architecture': arch_text, 'cpu_arch': cpu_arch_enum(arch_text, cpu_text),
        'gpu_model': flat(cell(r, 31)), 'gpu_cores': cores(cell(r, 32)),
        'gpu_clock_mhz': mhz(cell(r, 33)), 'ram_mb': ram_mb(cell(r, 34)),
        'screen_size_inch': num(cell(r, 35)), 'display_type': first(cell(r, 36)),
        'refresh_rate_hz': int(num(cell(r, 37))) if num(cell(r, 37)) else None,
        'screen_resolution_x': rx, 'screen_resolution_y': ry,
        'ppi': int(num(cell(r, 39))) if num(cell(r, 39)) else None,
        'aspect_ratio': aspect(raw(r, 40)),
        'screen_lens': flat(cell(r, 41)),
        'battery_capacity_mah': int(num(cell(r, 42))) if num(cell(r, 42)) else None,
        'cooling_solution': flat(cell(r, 43)),
        'ports': flat(cell(r, 49)),
        'storage_gb': sgb, 'storage_type': stype, 'storage_expandable': sexp,
        'microsd_type': microsd,
        'wifi_specs': wifi, 'bluetooth_specs': bt, 'other_connectivity': other_conn,
        'video_out': flat(cell(r, 52)),
        'audio_tech': audio_out,
        'audio_speakers': flat(cell(r, 54)),
        'has_headphone_jack': bool(audio_out and '3.5' in audio_out) or None,
        'haptics': tick(r, 55),
        'sensors': sensors,
        'has_microphone': bool(sensors and 'microphone' in sensors.lower()) or None,
        'gyro': bool(sensors and 'gyro' in sensors.lower()) or None,
        'performance_grade': stars(raw(r, 6)),
        'weight_g': int(num(cell(r, 61))) if num(cell(r, 61)) else None,
        'body_material': first(cell(r, 62)), 'available_colors': flat(cell(r, 63)),
        'width_mm': w, 'height_mm': h, 'depth_mm': dp,
        'touchscreen': bool(cell(r, 36) and 'touch' in str(cell(r, 36)).lower()) or None,
    }
    v = {k: x for k, x in v.items() if x is not None}

    # Buttons live on variant_input_profile, not console_variants.
    shape, place = parse_dpad(flat(cell(r, 44)))
    scount, stech, sclick, slayout = parse_analogs(flat(cell(r, 45)))
    btech, ttech, tlayout = parse_shoulder(flat(cell(r, 47)))
    ip = {
        'dpad_shape': shape, 'dpad_placement': place,
        'face_button_count': int(num(cell(r, 46))) if num(cell(r, 46)) else None,
        'stick_count': scount, 'stick_tech': stech,
        'stick_clicks': sclick if scount else None, 'stick_layout': slayout,
        'bumper_tech': btech, 'trigger_tech': ttech, 'trigger_layout': tlayout,
        'system_buttons_text': flat(cell(r, 48)),
        'has_gyro': v.get('gyro'),
        'input_notes': ' | '.join(filter(None, [flat(cell(r, 57)), flat(cell(r, 58)), flat(cell(r, 59))])) or None,
    }
    ip = {k: x for k, x in ip.items() if x is not None}
    if ip: v['input_profile'] = ip

    emu = {}
    for col, fields in EMU.items():
        rawv = cell(r, col)
        if not rawv: continue
        g = str(rawv).split('\n')[-1].strip()
        s = GRADE.get(g)
        if s:
            for f in fields: emu[f] = s
    limit = flat(cell(r, 79))
    if emu:
        if limit: emu['summary_text'] = limit
        v['emulation'] = emu

    # These columns show a label ("Retro Game Corps"); the URL lives in the cell's
    # hyperlink target, so reading the value alone yields no links at all.
    def linked(col):
        c = ws.cell(row=r, column=col)
        url = getattr(c.hyperlink, 'target', None) if c.hyperlink else None
        if not url and isinstance(c.value, str) and c.value.startswith('http'):
            url = c.value.strip()
        return (url, flat(c.value))

    links = []
    for i, col in enumerate([64, 65, 66, 67, 68]):
        u, label = linked(col)
        if u: links.append({'kind': 'video_review', 'url': u, 'label': label, 'sort_order': i})
    u, label = linked(69)
    if u: links.append({'kind': 'written_review', 'url': u, 'label': label, 'sort_order': 0})
    for i, col in enumerate([72, 73, 74, 75, 76]):
        u, label = linked(col)
        if u: links.append({'kind': 'vendor', 'url': u, 'label': label, 'sort_order': i})
    links = [{k: v for k, v in l.items() if v is not None} for l in links]

    splitlist = lambda s: [x.strip() for x in re.split(r'\n|;', str(s)) if x.strip()] if s else None

    console = {
        'name': disp, 'manufacturer': brand, 'slug': slug,
        'form_factor': (lambda ff: 'Clamshell' if ff and 'clamshell' in ff.lower()
                        else 'Vertical' if ff and 'vertical' in ff.lower()
                        else 'Horizontal' if ff and 'horizontal' in ff.lower() else 'Other')(cell(r, 4)),
        'device_category': CATEGORY,
        'status': 'draft',
        'release_status': 'upcoming' if (rel and rel > datetime.date.today().isoformat()) else 'released',
        'price_tier': flat(cell(r, 71)),
        'pros': splitlist(cell(r, 77)),
        'cons': splitlist(cell(r, 78)),
        'description': flat(cell(r, 80)),
        'variants': [v],
    }
    if links: console['links'] = links
    out.append({k: x for k, x in console.items() if x is not None})

json.dump(out, open(OUT, 'w'), indent=1, ensure_ascii=False)

slugs = [o['slug'] for o in out]
dupes = sorted({s for s in slugs if slugs.count(s) > 1})
filled = lambda key: sum(1 for o in out if o['variants'][0].get(key) is not None)
print(f"consoles: {len(out)} | unique slugs: {len(set(slugs))} | dupes: {dupes or 'none'}")
print(f"dropped (placeholder brand): {len(skipped)}")
print(f"with input_profile: {sum(1 for o in out if 'input_profile' in o['variants'][0])}")
print(f"with links: {sum(1 for o in out if 'links' in o)}")
print("field coverage:", {k: filled(k) for k in
      ['wifi_specs', 'bluetooth_specs', 'audio_speakers', 'haptics', 'ports',
       'screen_lens', 'storage_gb', 'microsd_type', 'performance_grade', 'cpu_arch']})

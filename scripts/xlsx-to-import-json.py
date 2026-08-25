import openpyxl, json, re

SRC='/root/.claude/uploads/454814d1-9ce8-55b6-9119-8737128283a8/c140fcaf-handhelds_v1.xlsx'
OUT='/tmp/claude-0/-home-user-The-Retro-Circuit/454814d1-9ce8-55b6-9119-8737128283a8/scratchpad/handhelds-import.json'

wb=openpyxl.load_workbook(SRC, data_only=True); ws=wb['Sheet1']

# Column -> emulation field, from the spreadsheet's own cell comments (verified, not guessed)
EMU={8:['gb_state','gbc_state'],9:['nes_state'],10:['genesis_state'],11:['gba_state'],
     12:['snes_state'],13:['ps1_state'],14:['nds_state'],15:['n64_state'],16:['dreamcast_state'],
     17:['psp_state'],18:['saturn_state'],19:['gamecube_state'],20:['wii_state'],
     21:['x3ds_state'],22:['ps2_state'],23:['wii_u'],24:['switch_state']}
GRADE={'A':'Perfect','A-':'Great','B':'Great','B-':'Playable','C':'Playable','D':'Struggles','F':'Unplayable'}

NULLISH={'❌','?','-','n/a','N/A','—',''}
def cell(r,c):
    v=ws.cell(row=r,column=c).value
    if v is None: return None
    s=str(v).strip()
    # The sheet uses ❌ / ? as "none/unknown" markers — never store those as values.
    if s in NULLISH: return None
    return s or None

def first(s):
    if not s: return None
    v=s.split('\n')[0].strip().rstrip(',;').strip()
    return v or None

def flat(s):
    if not s: return None
    v=' '.join(x.strip() for x in s.split('\n') if x.strip()).strip().rstrip(',;').strip()
    v=re.sub(r'\s*,\s*,', ',', v)
    return v or None

def num(s):
    if not s: return None
    m=re.search(r'(\d+(?:\.\d+)?)', s.replace(',',''))
    return float(m.group(1)) if m else None

def mhz(s):
    """Take the highest clock mentioned, converting GHz->MHz."""
    if not s or s in ('?','❌'): return None
    best=None
    for val,unit in re.findall(r'(\d+(?:\.\d+)?)\s*(GHz|MHz)', s, re.I):
        v=float(val)*(1000 if unit.lower()=='ghz' else 1)
        best=v if best is None else max(best,v)
    return int(round(best)) if best else None

def ram_mb(s):
    if not s: return None
    m=re.search(r'(\d+(?:\.\d+)?)\s*(GB|MB|KB)', s, re.I)
    if not m: return None
    v=float(m.group(1)); u=m.group(2).upper()
    return int(v*1024) if u=='GB' else int(v) if u=='MB' else 0

def cores(s):
    if not s: return None
    words={'single':1,'dual':2,'triple':3,'quad':4,'hexa':6,'octa':8}
    low=s.lower()
    for w,n in words.items():
        if w in low: return n
    m=re.search(r'(\d+)\s*core', low)
    return int(m.group(1)) if m else None

def resolution(s):
    if not s: return (None,None)
    m=re.search(r'(\d+)\s*[×x]\s*(\d+)', s)
    return (int(m.group(1)), int(m.group(2))) if m else (None,None)

def dims(s):
    if not s: return (None,None,None)
    m=re.findall(r'(\d+(?:\.\d+)?)\s*mm', s)
    return (float(m[0]),float(m[1]),float(m[2])) if len(m)>=3 else (None,None,None)

def release(s):
    """'2004 / 11' -> (date, precision)."""
    if not s: return (None,None)
    m=re.match(r'(\d{4})\s*/\s*(\d{1,2})', s)
    if m: return (f"{m.group(1)}-{int(m.group(2)):02d}-01",'month')
    m=re.match(r'(\d{4})', s)
    return (f"{m.group(1)}-01-01",'year') if m else (None,None)

def os_bits(s):
    """Derive structured os_family/os_version from the stock-OS string."""
    if not s: return (None,None)
    low=s.lower()
    fam=('android' if 'android' in low else
         'windows' if 'windows' in low else
         'steamos' if 'steamos' in low else
         'linux' if 'linux' in low else 'proprietary')
    m=re.search(r'(\d+(?:\.\d+)*)', s)
    return (fam, m.group(1) if m else None)

def form_factor(s):
    if not s: return None
    low=s.lower()
    if 'clamshell' in low: return 'Clamshell'
    if 'portrait' in low: return 'Vertical'
    if 'landscape' in low: return 'Horizontal'
    return 'Other'

def slugify(s):
    return re.sub(r'(^-|-$)','', re.sub(r'[^a-z0-9]+','-', s.lower().strip()))

MFG={'Nintendo':'Nintendo','Sony':'Sony','Sony Ericsson':'Sony Ericsson'}

out=[]
for r in range(2, ws.max_row+1):
    name=flat(cell(r,1))
    if not name: continue
    brand_raw=flat(cell(r,2)) or ''
    brand = 'Panic' if 'panic' in brand_raw.lower() else brand_raw
    rel,prec = release(cell(r,3))
    os_raw = first(cell(r,5))
    fam,ver = os_bits(os_raw)
    rx,ry = resolution(cell(r,35))
    w,h,d = dims(cell(r,57))
    price = cell(r,66)
    price_val = int(num(price)) if price and 'discontinu' not in price.lower() else None
    vid = cell(r,49)
    storage_raw = flat(cell(r,46))

    emu={}
    for col,fields in EMU.items():
        raw=cell(r,col)
        if not raw: continue
        g=raw.split('\n')[-1].strip()
        state=GRADE.get(g)
        if state:
            for f in fields: emu[f]=state

    variant={
        'variant_name':'Base','is_default':True,
        'release_date':rel,'release_date_precision':prec,
        'price_launch_usd':price_val,
        'os':os_raw,'os_family':fam,'os_version':ver,
        'soc':first(cell(r,25)),
        'cpu_model':flat(cell(r,26)),'cpu_cores':cores(cell(r,27)),
        'cpu_clock_max_mhz':mhz(cell(r,28)),'cpu_architecture':first(cell(r,29)),
        'gpu_model':flat(cell(r,30)),'gpu_clock_mhz':mhz(cell(r,31)),
        'ram_mb':ram_mb(cell(r,32)),
        'screen_size_inch':num(cell(r,33)),'display_type':first(cell(r,34)),
        'screen_resolution_x':rx,'screen_resolution_y':ry,
        'aspect_ratio':first(cell(r,36)),
        'battery_capacity_mah':int(num(cell(r,38))) if num(cell(r,38)) else None,
        'cooling_solution':flat(cell(r,39)),
        'storage_type':storage_raw,
        'storage_expandable': bool(storage_raw and ('external' in storage_raw.lower() or 'microsd' in storage_raw.lower())),
        'video_out': None if (not vid or vid=='❌') else flat(vid),
        'weight_g':int(num(cell(r,58))) if num(cell(r,58)) else None,
        'available_colors':flat(cell(r,59)),
        'width_mm':w,'height_mm':h,'depth_mm':d,
        'touchscreen': bool(cell(r,34) and 'touch' in str(cell(r,34)).lower()),
    }
    variant={k:v for k,v in variant.items() if v is not None}
    if emu: variant['emulation']=emu

    out.append({k:v for k,v in {
        'name':name,'manufacturer':brand,
        'slug':slugify(f"{brand} {name}"),
        'form_factor':form_factor(cell(r,4)),
        'device_category':'legacy',
        'status':'draft',
        'variants':[variant],
    }.items() if v is not None})

json.dump(out, open(OUT,'w'), indent=2, ensure_ascii=False)
print(f"wrote {len(out)} consoles -> {OUT}")
brands={}
for o in out: brands[o['manufacturer']]=brands.get(o['manufacturer'],0)+1
print("brands:", brands)
